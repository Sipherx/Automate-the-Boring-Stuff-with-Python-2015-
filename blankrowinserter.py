#!usr/bin/env python

from sys import argv
import openpyxl
from openpyxl.cell import get_column_letter
script, n, m, filename = argv

wb = openpyxl.load_workbook(filename)
ws = wb.active
ws1 = wb.create_sheet(title = 'copy')

for row in range(1, int(n)+1):
	for col in range(1, ws.max_column+1):
		ws1[get_column_letter(col) + str(row)] = ws[get_column_letter(col) + str(row)].value
		
for row1 in range(1, ws.max_row-int(m)+2):
	for col1 in range(1, ws.max_column+1):
		ws1[get_column_letter(col1) + str(row1+int(n)+int(m))] = ws[get_column_letter(col1) + str(row1+int(n))].value
	
wb.save(filename +'_copy.xlsx')		


