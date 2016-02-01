#!usr/bin/env python
# multiplication table.py - creates the table with user defined number 
# in command line
import openpyxl, sys
from openpyxl.styles import Font
from openpyxl.cell import get_column_letter
from openpyxl.compat import range

number = int(sys.argv[1])

wb = openpyxl.Workbook()
ws1 = wb.active

bf = Font(bold=True)
	
for row in range(1, number+1):
	# set column to bold
	ColumnLabel = ws1['A' + str(row)]
	ColumnLabel.font = bf
	# set row to bold
	RowLabel = ws1[get_column_letter(row) + str(1)]
	RowLabel.font = bf
	
	for col in range(1, number+1):
		value = ws1.cell(column = col, row = row, value = row*col)
	
wb.save('multitable.xlsx')


 
