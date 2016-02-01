#!usr/bin/env python

import openpyxl, sys
from openpyxl.styles import Font, Style
from openpyxl.cell import get_column_letter

number = int(sys.argv[1])

wb = openpyxl.Workbook()
sheet = wb.active

row = 1
column = 1

bf = Font(bold=True)


	
for cell in range(1, number+1):
	columnLabel = sheet['A' + str(cell)]
	columnLabel.bf
	#sheet[get_column_letter(cell) + str(1)]
	sheet[get_column_letter(cell) + str(1)] = cell
	
wb.save('multitable.xlsx')


 
